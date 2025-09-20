import json
import os
import pandas as pd
from openpyxl import load_workbook
import glob

def extract_header_hierarchy(ws, header_rows=3):
    """Extract header hierarchy from worksheet starting from row 1"""
    # Build a lookup for merged cell ranges
    merged_lookup = {}
    for merged in ws.merged_cells.ranges:
        top_left = ws.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                merged_lookup[(row, col)] = top_left
    
    # Collect headers, replacing merged cells on read
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=header_rows):
        headers.append([
            merged_lookup.get((cell.row, cell.column), cell.value)
            for cell in row
        ])
    
    # Transpose to work column-wise
    headers_t = list(zip(*headers))
    
    # Create flattened column names
    flattened_headers = []
    for col in headers_t:
        # Remove None and collapse consecutive duplicates
        levels = []
        prev = None
        for h in col:
            if h is not None and h != prev:
                levels.append(str(h).strip())
            prev = h
        
        # Join levels with underscore to create flattened column name
        if levels:
            flattened_name = '_'.join(levels)
            flattened_headers.append(flattened_name)
        else:
            flattened_headers.append(f"Column_{len(flattened_headers) + 1}")
    
    return flattened_headers

def extract_data_from_excel(file_path, sheet_name=0, skip_rows=7, header_rows=3):
    """Extract data from Excel file after skipping specified rows"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]
    
    # Get the maximum row and column with data
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Extract data starting from (skip_rows + header_rows + 1)
    data_start_row = skip_rows + header_rows + 1
    data = []
    
    for row in ws.iter_rows(min_row=data_start_row, max_row=max_row, 
                           min_col=1, max_col=max_col, values_only=True):
        # Only add rows that have at least one non-None value
        if any(cell is not None for cell in row):
            data.append(list(row))
    
    wb.close()
    return data

def get_headers_from_excel(file_path, sheet_name=0, skip_rows=7, header_rows=3):
    """Get flattened headers from Excel file after skipping specified rows"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]
    
    # Create a new temporary worksheet with headers moved to top
    # We need to extract the header rows after skipping the first 7 rows
    header_data = []
    for row in ws.iter_rows(min_row=skip_rows + 1, max_row=skip_rows + header_rows):
        header_data.append([cell.value for cell in row])
    
    # Create temporary worksheet-like structure for header extraction
    # We need to handle merged cells in the header area
    merged_lookup = {}
    for merged in ws.merged_cells.ranges:
        # Only consider merged cells in our header area
        if (skip_rows + 1 <= merged.min_row <= skip_rows + header_rows):
            top_left = ws.cell(merged.min_row, merged.min_col).value
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    # Adjust row number to be relative to our header start
                    relative_row = row - skip_rows
                    merged_lookup[(relative_row, col)] = top_left
    
    # Process headers with merged cell values
    processed_headers = []
    for i, row_data in enumerate(header_data):
        processed_row = []
        for j, cell_value in enumerate(row_data):
            # Check if this cell is part of a merged range
            merged_value = merged_lookup.get((i + 1, j + 1), cell_value)
            processed_row.append(merged_value)
        processed_headers.append(processed_row)
    
    # Transpose to work column-wise
    headers_t = list(zip(*processed_headers))
    
    # Create flattened column names
    flattened_headers = []
    for col in headers_t:
        # Remove None and collapse consecutive duplicates
        levels = []
        prev = None
        for h in col:
            if h is not None and h != prev:
                levels.append(str(h).strip())
            prev = h
        
        # Join levels with underscore to create flattened column name
        if levels:
            flattened_name = '_'.join(levels)
            flattened_headers.append(flattened_name)
        else:
            flattened_headers.append(f"Column_{len(flattened_headers) + 1}")
    
    wb.close()
    return flattened_headers

def process_excel_to_csv(directory_path, output_directory=None):
    """Process all Excel files in directory and convert to CSV"""
    
    # Set output directory
    if output_directory is None:
        output_directory = os.path.join(directory_path, "csv_output")
    
    # Create output directory if it doesn't exist
    os.makedirs(output_directory, exist_ok=True)
    
    # Find all Excel files
    excel_files = glob.glob(os.path.join(directory_path, "*.xlsx")) + \
                 glob.glob(os.path.join(directory_path, "*.xls"))
    
    if not excel_files:
        print("No Excel files found in the specified directory.")
        return
    
    print(f"Found {len(excel_files)} Excel files to process.")
    
    for file_path in excel_files:
        try:
            print(f"\nProcessing: {os.path.basename(file_path)}")
            
            # Extract headers
            headers = get_headers_from_excel(file_path, skip_rows=7, header_rows=3)
            print(f"Extracted {len(headers)} headers")
            
            # Extract data
            data = extract_data_from_excel(file_path, skip_rows=7, header_rows=3)
            print(f"Extracted {len(data)} data rows")
            
            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            # Generate output filename
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_file = os.path.join(output_directory, f"{base_name}.csv")
            
            # Save to CSV
            df.to_csv(output_file, index=False, encoding='utf-8')
            print(f"Saved: {output_file}")
            
            # Also save headers structure as JSON for reference
            json_file = os.path.join(output_directory, f"{base_name}_headers.json")
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(headers, f, indent=2, ensure_ascii=False)
            
        except Exception as e:
            print(f"Error processing {os.path.basename(file_path)}: {str(e)}")
            continue
    
    print(f"\nProcessing complete. Output files saved in: {output_directory}")

def main():
    # Configuration
    directory_path = input("Enter the directory path containing Excel files: ").strip()
    
    if not os.path.exists(directory_path):
        print("Directory does not exist!")
        return
    
    # Optional: specify output directory
    output_dir = input("Enter output directory (press Enter for default 'csv_output'): ").strip()
    if not output_dir:
        output_dir = None
    
    # Process files
    process_excel_to_csv(directory_path, output_dir)

if __name__ == "__main__":
    main()